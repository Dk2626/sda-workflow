#!/usr/bin/env python3
"""
Generate combined Excel workbook:
- Sheet 1: Test Scripts (Functional Test Cases) - Deliverable 6
- Sheet 2: Traceability Matrix - Deliverable 7
- Sheet 3: AI Tool Usage Matrix (bonus)
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

# ============================================================================
# Common styles
# ============================================================================
hdr_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
hdr_fill = PatternFill('solid', start_color='0A2540')
hdr_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

cell_font = Font(name='Arial', size=10)
cell_align = Alignment(vertical='top', wrap_text=True)

thin = Side(style='thin', color='CCCCCC')
border = Border(left=thin, right=thin, top=thin, bottom=thin)

def apply_header(ws, row=1):
    for cell in ws[row]:
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = hdr_align
        cell.border = border
    ws.row_dimensions[row].height = 30

def apply_data_style(ws, start_row=2):
    for row in ws.iter_rows(min_row=start_row, max_row=ws.max_row):
        for cell in row:
            cell.font = cell_font
            cell.alignment = cell_align
            cell.border = border

# ============================================================================
# SHEET 1: Test Scripts (Functional Test Cases)
# ============================================================================
ws1 = wb.active
ws1.title = "Test Scripts"

test_cases = [
    # (TC_ID, Module, Title, Pre-conditions, Steps, Expected Result, Type, FR/US Map, Priority)
    # ----- Authentication -----
    ('TC-001', 'Auth', 'Valid login', 'User alice exists with role REQUESTOR',
     '1. Open /login\n2. Enter username=alice, password=demo123\n3. Submit',
     'Redirected to /. Session cookie set.', 'Positive', 'NFR-003', 'High'),
    ('TC-002', 'Auth', 'Invalid password', 'User alice exists',
     '1. Open /login\n2. Enter username=alice, password=wrong\n3. Submit',
     '401 with "Invalid credentials" error.', 'Negative', 'NFR-003', 'High'),
    ('TC-003', 'Auth', 'Empty credentials', '—',
     '1. POST /api/auth/login with empty body',
     '400 with "Username and password required".', 'Negative', 'NFR-003', 'Med'),
    ('TC-004', 'Auth', 'Logout clears session', 'User logged in',
     '1. POST /api/auth/logout\n2. GET /api/me',
     '401 Not authenticated.', 'Positive', 'NFR-003', 'High'),

    # ----- Deviation Creation (FR-001, FR-005) -----
    ('TC-010', 'Create', 'Create valid billing-credit draft', 'Logged in as alice',
     '1. POST /api/deviations with customer_id, deviation_type=DT_BILL_CR, action, amount=2000, 50+ char justification',
     '201 returned with reference SDA-YYYY-NNNNN and state=DRAFT.', 'Positive', 'FR-001, US-01', 'High'),
    ('TC-011', 'Create', 'Missing customer rejected', 'Logged in as alice',
     '1. POST /api/deviations without customer_id',
     '400 "Missing required fields".', 'Negative', 'FR-005', 'High'),
    ('TC-012', 'Create', 'Justification under 50 chars rejected', 'Logged in as alice',
     '1. POST /api/deviations with justification="too short"',
     '400 "Justification must be 50-4000 characters".', 'Negative', 'FR-005', 'High'),
    ('TC-013', 'Create', 'Negative amount rejected', 'Logged in as alice',
     '1. POST /api/deviations with amount=-100',
     '400 "Amount must be non-negative".', 'Negative', 'FR-005', 'High'),
    ('TC-014', 'Create', 'Approver cannot create deviations', 'Logged in as bob (L1_APPROVER)',
     '1. POST /api/deviations',
     '403 Forbidden.', 'Negative', 'FR-018', 'High'),

    # ----- AI Justification (FR-002, FR-015) -----
    ('TC-020', 'AI Just.', 'Generate justification text', 'Logged in as alice; customer selected',
     '1. POST /api/ai/justification with valid context',
     '200 with non-empty justification text. Invocation logged to ai_invocations table.', 'Positive', 'FR-002, US-02', 'High'),
    ('TC-021', 'AI Just.', 'Missing context rejected', 'Logged in as alice',
     '1. POST /api/ai/justification with empty body',
     '400 "Missing required context".', 'Negative', 'FR-002', 'Med'),
    ('TC-022', 'AI Just.', 'AI fallback on API error', 'OPENAI_API_KEY invalid',
     '1. POST /api/ai/justification',
     '200 with fallback justification text. ai_invocations row has model=fallback-mock.', 'Boundary', 'FR-015', 'High'),

    # ----- AI Risk Scoring (FR-003) -----
    ('TC-030', 'AI Risk', 'Risk score on submit', 'Draft deviation exists for alice',
     '1. POST /api/deviations/:id/submit',
     'State=UNDER_REVIEW. ai_risk_score set to LOW/MEDIUM/HIGH. Factors stored as JSON.', 'Positive', 'FR-003, US-04', 'High'),
    ('TC-031', 'AI Risk', 'High amount yields higher risk band', 'Draft with amount=20000',
     '1. Submit',
     'ai_risk_score is MEDIUM or HIGH (depending on customer tier).', 'Positive', 'FR-003', 'High'),
    ('TC-032', 'AI Risk', 'Risk score JSON valid', 'After submit',
     '1. GET /api/deviations/:id',
     'ai_risk_factors is valid JSON array, each item has factor/weight/value.', 'Positive', 'FR-015', 'Med'),

    # ----- Workflow / Approval Routing (FR-004, FR-006) -----
    ('TC-040', 'Workflow', 'Submit routes to L1 first', 'Draft submitted',
     '1. Submit\n2. Inspect current_approver_role',
     'current_approver_role = L1_APPROVER, sla_deadline ~8h ahead.', 'Positive', 'FR-004', 'High'),
    ('TC-041', 'Workflow', 'L1 approval routes to Finance for DT_BILL_CR', 'DT_BILL_CR UNDER_REVIEW; logged as bob',
     '1. POST /api/deviations/:id/approve with comment',
     'state=UNDER_REVIEW, current_approver_role=FINANCE_APPROVER, action recorded.', 'Positive', 'FR-004, FR-006', 'High'),
    ('TC-042', 'Workflow', 'Compliance required when amount > 5000', 'DT_BILL_CR amount=6000',
     '1. Approve as L1\n2. Approve as Finance\n3. Inspect next_role',
     'next_role = COMPLIANCE_APPROVER. State not FINAL_APPROVED yet.', 'Positive', 'FR-004, FR-009', 'High'),
    ('TC-043', 'Workflow', 'Compliance not required when amount = 3000', 'DT_BILL_CR amount=3000',
     '1. Approve as L1\n2. Approve as Finance',
     'state = FINAL_APPROVED.', 'Positive', 'FR-004', 'High'),
    ('TC-044', 'Workflow', 'KYC always needs compliance', 'DT_KYC_DF submitted',
     '1. Approve as L1\n2. Inspect next role',
     'next_role = COMPLIANCE_APPROVER.', 'Positive', 'FR-004, FR-009', 'High'),
    ('TC-045', 'Workflow', 'Wrong role cannot approve', 'UNDER_REVIEW awaiting L1, logged as finn',
     '1. POST /api/deviations/:id/approve',
     '403 "awaiting L1_APPROVER, not FINANCE_APPROVER".', 'Negative', 'FR-018', 'High'),
    ('TC-046', 'Workflow', 'Cannot approve own request', 'alice owns request; alice somehow as approver',
     '1. Approve own',
     '403 "Requestor cannot approve own request".', 'Negative', 'FR-018', 'High'),
    ('TC-047', 'Workflow', 'Comment required on approve', 'L1 logged in',
     '1. POST approve with empty comments',
     '400 "comment of at least 5 characters required".', 'Negative', 'FR-006', 'Med'),

    # ----- Reject & Veto (FR-006, FR-009) -----
    ('TC-050', 'Reject', 'L1 reject moves state to REJECTED', 'UNDER_REVIEW',
     '1. Reject with comment',
     'state=REJECTED, current_approver_role=NULL, audit row REJECT.', 'Positive', 'FR-006', 'High'),
    ('TC-051', 'Reject', 'Compliance veto creates COMPLIANCE_VETO audit', 'APPROVED_L2 awaiting Compliance, KYC type',
     '1. Reject as carol',
     'state=REJECTED, audit event_type=COMPLIANCE_VETO.', 'Positive', 'FR-009, US-06', 'High'),
    ('TC-052', 'Reject', 'Cannot reject DRAFT', 'DRAFT state',
     '1. POST reject',
     '400 "Cannot reject from state DRAFT".', 'Negative', 'FR-006', 'Med'),

    # ----- Info request -----
    ('TC-060', 'Info Req.', 'Request info transitions to INFO_REQUESTED', 'UNDER_REVIEW',
     '1. POST /request-info',
     'state=INFO_REQUESTED.', 'Positive', 'FR-006', 'Med'),
    ('TC-061', 'Info Req.', 'Cannot request info on FINAL_APPROVED', 'FINAL_APPROVED state',
     '1. POST /request-info',
     '400.', 'Negative', 'FR-006', 'Low'),

    # ----- Risk Override (FR-013) -----
    ('TC-070', 'Override', 'Valid risk override saved', 'UNDER_REVIEW; logged in L1',
     '1. POST /override-risk with new_score=HIGH, reason 20+ chars',
     'ai_risk_score updated. approval_actions row OVERRIDE_RISK with reason.', 'Positive', 'FR-013, US-08', 'High'),
    ('TC-071', 'Override', 'Short reason rejected', 'Logged in approver',
     '1. POST /override-risk with reason="short"',
     '400 "reason of at least 20 characters".', 'Negative', 'FR-013', 'High'),
    ('TC-072', 'Override', 'Invalid score rejected', 'Logged in approver',
     '1. POST /override-risk with new_score=CRITICAL',
     '400 "new_score must be LOW/MEDIUM/HIGH".', 'Negative', 'FR-013', 'Med'),

    # ----- AI Context Summary -----
    ('TC-080', 'AI Context', 'Generate summary first time', 'UNDER_REVIEW; logged L1',
     '1. POST /api/ai/context-summary',
     '200 with summary text. Stored on deviation row.', 'Positive', 'US-03, FR-015', 'High'),
    ('TC-081', 'AI Context', 'Subsequent call returns cached', 'Summary already generated',
     '1. POST /api/ai/context-summary again',
     '200 with cached=true.', 'Positive', 'FR-015', 'Med'),

    # ----- AI Policy -----
    ('TC-090', 'AI Policy', 'Citations returned for known type', 'UNDER_REVIEW DT_BILL_CR',
     '1. POST /api/ai/policy',
     '200 with at least 1 citation, fields policy_code/title/snippet/relevance.', 'Positive', 'US-07, FR-011', 'High'),
    ('TC-091', 'AI Policy', 'Citations filtered to deviation type', 'DT_KYC_DF',
     '1. POST /api/ai/policy',
     'Citations relate to KYC policy POL-KYC-004.', 'Positive', 'FR-011', 'Med'),

    # ----- Audit (FR-008, FR-014) -----
    ('TC-100', 'Audit', 'Every state change logged', 'New deviation submitted, approved L1',
     '1. GET /api/deviations/:id/audit',
     'audit_log has CREATE_DRAFT, SUBMIT, APPROVE entries.', 'Positive', 'FR-008, US-09', 'High'),
    ('TC-101', 'Audit', 'AI invocations linked to deviation', 'Submit triggered scoring',
     '1. GET audit',
     'ai_invocations array non-empty, agent_type RISK_SCORING present.', 'Positive', 'FR-015', 'High'),
    ('TC-102', 'Audit', 'No audit edit endpoint exists', '—',
     '1. Try PUT/DELETE on /api/audit',
     '404 Method not allowed. Append-only confirmed.', 'Security', 'NFR-004', 'High'),
    ('TC-103', 'Audit', 'CSV export downloadable', 'Logged in',
     '1. GET /api/reports/export',
     '200 with text/csv content-type and attachment header.', 'Positive', 'FR-014', 'Med'),

    # ----- Execute (FR-017) -----
    ('TC-110', 'Execute', 'Execute moves FINAL_APPROVED to EXECUTED', 'FINAL_APPROVED; alice owner',
     '1. POST /execute',
     'state=EXECUTED. Audit EXECUTE row.', 'Positive', 'FR-016', 'High'),
    ('TC-111', 'Execute', 'Cannot execute non-final', 'state=APPROVED_L1',
     '1. POST /execute',
     '400 "Cannot execute from state APPROVED_L1".', 'Negative', 'FR-017', 'Med'),
    ('TC-112', 'Execute', 'Non-owner cannot execute', 'FINAL_APPROVED; logged in bob',
     '1. POST /execute',
     '403 Forbidden.', 'Negative', 'FR-018', 'High'),

    # ----- Visibility / RBAC -----
    ('TC-120', 'RBAC', 'Requestor sees only own deviations', 'alice has 1, bob has 0',
     '1. GET /api/deviations as alice',
     'List contains only alice\'s deviations.', 'Security', 'FR-018', 'High'),
    ('TC-121', 'RBAC', 'Approver sees all deviations', 'Multiple exist',
     '1. GET /api/deviations as bob',
     'List contains all deviations.', 'Positive', 'FR-018', 'High'),
    ('TC-122', 'RBAC', 'Auditor read-only', 'logged in as aud',
     '1. POST /api/deviations',
     '403 Forbidden.', 'Security', 'FR-018', 'High'),

    # ----- UI -----
    ('TC-130', 'UI', 'Login page renders', '—',
     '1. Open /login',
     'Form with username/password fields visible.', 'UI', 'NFR-006', 'Low'),
    ('TC-131', 'UI', 'Dashboard tiles populate', 'Logged in, data exists',
     '1. Open /',
     'Tiles show counts; Recent Deviations table visible.', 'UI', 'FR-010, US-12', 'Med'),
    ('TC-132', 'UI', 'Deviation form AI button generates text', 'Logged as alice, customer selected',
     '1. Click "Generate with AI"\n2. Wait <8s',
     'Justification textarea populated with AI text.', 'UI', 'FR-002, US-02', 'High'),
    ('TC-133', 'UI', 'Approver workbench shows risk badge', 'UNDER_REVIEW with risk score',
     '1. Open /deviations/:id as L1',
     'Risk badge LOW/MED/HIGH visible with correct color.', 'UI', 'US-04', 'High'),
    ('TC-134', 'UI', 'Audit page renders trail', 'Auditor logged in',
     '1. Open /audit?id=1',
     'Audit events and AI invocations listed.', 'UI', 'US-09', 'Med'),

    # ----- Performance -----
    ('TC-140', 'Perf.', 'AI call completes <8s', 'OPENAI_API_KEY set',
     '1. POST /api/ai/justification\n2. Measure',
     'Response time <8000ms (per NFR-001).', 'Performance', 'NFR-001', 'Med'),
    ('TC-141', 'Perf.', 'Deviation list <1.5s p95', '50+ deviations seeded',
     '1. GET /api/deviations 20 times',
     '95th percentile under 1.5s.', 'Performance', 'NFR-001', 'Med'),

    # ----- Reference uniqueness -----
    ('TC-150', 'Data', 'Reference format correct', '—',
     '1. Create 3 deviations\n2. Inspect references',
     'All follow pattern SDA-YYYY-NNNNN and are unique.', 'Positive', 'FR-008', 'High'),
    ('TC-151', 'Data', 'PII redacted in AI prompts', 'Customer name set',
     '1. Trigger justification AI\n2. Inspect ai_invocations.prompt',
     'Customer name absent or replaced with [CUSTOMER].', 'Security', 'NFR-003', 'High'),
]

ws1.append(['TC ID', 'Module', 'Title', 'Pre-conditions', 'Test Steps', 'Expected Result', 'Type', 'Maps to FR/US', 'Priority'])
for tc in test_cases:
    ws1.append(list(tc))

# Column widths
widths1 = {'A': 10, 'B': 12, 'C': 38, 'D': 30, 'E': 50, 'F': 50, 'G': 12, 'H': 18, 'I': 10}
for col, w in widths1.items():
    ws1.column_dimensions[col].width = w

apply_header(ws1)
apply_data_style(ws1)

# Freeze top row
ws1.freeze_panes = 'A2'

# ============================================================================
# SHEET 2: Traceability Matrix
# ============================================================================
ws2 = wb.create_sheet("Traceability Matrix")

matrix = [
    # (Req ID, Requirement Summary, User Story, Spec/DSL Element, Code File(s), Test Case IDs)
    ('FR-001', 'Create deviation with full payload',                   'US-01', 'specs/00_main.yaml > entities.DeviationRequest; specs/01_user_stories.yaml > US-01', 'app/src/app/api/deviations/route.js (POST)', 'TC-010, TC-011, TC-014'),
    ('FR-002', 'Invoke AI Justification agent on demand',              'US-02', 'specs/00_main.yaml > ai_agents.justification_agent',                                'app/src/app/api/ai/justification/route.js; app/src/lib/ai-agents.js#generateJustification', 'TC-020, TC-021, TC-022, TC-132'),
    ('FR-003', 'Compute AI risk score on submission',                  'US-04', 'specs/00_main.yaml > ai_agents.risk_scoring_agent',                                  'app/src/app/api/deviations/[id]/submit/route.js; app/src/lib/ai-agents.js#scoreRisk', 'TC-030, TC-031, TC-032'),
    ('FR-004', 'Route via approval chain per deviation type',          'US-01, US-06', 'specs/00_main.yaml > approval_chains',                                            'app/src/lib/workflow.js (chainFor, nextApproverRole)', 'TC-040, TC-041, TC-042, TC-043, TC-044'),
    ('FR-005', 'Block submit if mandatory fields/attachments missing', 'US-01', 'specs/00_main.yaml > validation',                                                    'app/src/app/api/deviations/route.js (POST validation block)', 'TC-011, TC-012, TC-013'),
    ('FR-006', 'Approve/Reject/Request-Info with comment',             'US-06', 'specs/00_main.yaml > state_machine.transitions',                                    'app/src/app/api/deviations/[id]/approve|reject|request-info/route.js', 'TC-041, TC-047, TC-050, TC-060'),
    ('FR-007', 'SLA timers + auto-escalation',                         'US-10', 'specs/00_main.yaml > sla',                                                          'app/src/lib/workflow.js#slaDeadline (escalation worker out-of-scope for demo)', 'TC-040'),
    ('FR-008', 'Audit log on every state transition and AI call',      'US-09', 'specs/00_main.yaml > entities.AuditLog, entities.AiInvocation',                     'app/src/lib/workflow.js#logAudit; app/src/lib/ai-agents.js#logInvocation', 'TC-100, TC-101, TC-102, TC-150'),
    ('FR-009', 'Compliance veto power on any high-risk/KYC deviation', 'US-06', 'specs/00_main.yaml > approval_chains + state_machine',                                'app/src/app/api/deviations/[id]/reject/route.js (COMPLIANCE_VETO branch)', 'TC-051, TC-042'),
    ('FR-010', 'Dashboard pending/approved/rejected/overdue',          'US-12', 'specs/00_main.yaml > api.GET /reports/dashboard',                                   'app/src/app/api/reports/dashboard/route.js; app/src/app/page.js', 'TC-131'),
    ('FR-011', 'AI policy citations on demand',                        'US-07', 'specs/00_main.yaml > ai_agents.policy_agent',                                       'app/src/app/api/ai/policy/route.js; app/src/lib/ai-agents.js#interpretPolicy', 'TC-090, TC-091'),
    ('FR-012', 'Notifications on state changes',                       'US-11', 'specs/00_main.yaml > notifications',                                                'app/src/lib/workflow.js (event hooks; SMTP integration mocked)', '—'),
    ('FR-013', 'Override AI risk with mandatory reason',               'US-08', 'specs/00_main.yaml > state_machine.OVERRIDE_RISK',                                   'app/src/app/api/deviations/[id]/override-risk/route.js', 'TC-070, TC-071, TC-072'),
    ('FR-014', 'Audit-ready CSV export',                               'US-09', 'specs/00_main.yaml > api.GET /reports/export',                                      'app/src/app/api/reports/export/route.js', 'TC-103'),
    ('FR-015', 'AI explainability metadata persisted',                 'US-09', 'specs/00_main.yaml > entities.AiInvocation',                                        'app/src/lib/ai-agents.js#logInvocation', 'TC-022, TC-032, TC-080, TC-101'),
    ('FR-016', 'Requestor can withdraw before FINAL_APPROVED',         '—',     'specs/00_main.yaml > state_machine.WITHDRAW',                                       'app/src/lib/workflow.js (WITHDRAW action — endpoint stub)', 'TC-110'),
    ('FR-017', 'Expire FINAL_APPROVED after window',                   '—',     'specs/00_main.yaml > state_machine.EXPIRE',                                          'Background worker stub (out-of-scope for demo)', 'TC-111'),
    ('FR-018', 'RBAC with segregation of duties',                      'US-01', 'specs/00_main.yaml > security.rbac; validation.segregation_of_duties',              'app/src/lib/api-helpers.js#requireUser; approve route SoD check', 'TC-014, TC-045, TC-046, TC-112, TC-120, TC-121, TC-122'),
    ('NFR-001', 'Performance: UI <1.5s, AI <8s',                       '—',     'specs/00_main.yaml > nfr.performance',                                              'In-memory ops + SQLite; no slow paths', 'TC-140, TC-141'),
    ('NFR-002', 'Availability 99.5%',                                  '—',     '—',                                                                                  'Infrastructure scope (out-of-app)', '—'),
    ('NFR-003', 'Security: OWASP + encryption + RBAC',                 '—',     'specs/00_main.yaml > security',                                                     'app/src/lib/auth.js (bcrypt, session); PII redaction in ai-agents.js', 'TC-001, TC-002, TC-003, TC-004, TC-151'),
    ('NFR-004', 'Audit retention 7+ years, append-only',               'US-09', 'specs/00_main.yaml > security.audit_retention_years',                              'app/src/lib/workflow.js#logAudit (no update/delete endpoints)', 'TC-102'),
    ('NFR-005', 'Scalability 10k requests/month',                      '—',     '—',                                                                                  'SQLite OK for demo; PG migration noted in db/database_design.md', '—'),
    ('NFR-006', 'Usability: approver review <2min',                    '—',     'specs/00_main.yaml > nfr.usability',                                                'Approver workbench design app/src/app/deviations/[id]/page.js', 'TC-130, TC-133, TC-134'),
    ('NFR-007', 'Observability: structured logs, traces, metrics',     '—',     '—',                                                                                  'Console logging in dev; full observability stack out-of-scope', '—'),
]

ws2.append(['Requirement ID', 'Requirement Summary', 'User Story', 'Spec/DSL Element', 'Code File(s)', 'Test Case IDs'])
for row in matrix:
    ws2.append(list(row))

widths2 = {'A': 12, 'B': 38, 'C': 14, 'D': 50, 'E': 60, 'F': 30}
for col, w in widths2.items():
    ws2.column_dimensions[col].width = w

apply_header(ws2)
apply_data_style(ws2)
ws2.freeze_panes = 'A2'

# ============================================================================
# SHEET 3: AI Tool Usage Matrix (bonus deliverable)
# ============================================================================
ws3 = wb.create_sheet("AI Tool Usage")

ai_usage = [
    # SDLC phase, Task, AI tool used, How AI was used, Output artifact
    ('Requirements', 'Drafting Functional Requirements Document',
     'Claude (Anthropic)',
     'Provided use-case PDF; Claude analyzed scope, generated 18 functional requirements, 12 user stories, 7 NFRs across all sections of FRD with consistent IDs and traceability.',
     'docs/01_FRD.html'),

    ('Requirements', 'User story decomposition (US-01..US-12)',
     'Claude',
     'Asked Claude to derive personas, acceptance criteria, and FR mappings for each user story. Used Given/When/Then style.',
     'specs/01_user_stories.yaml'),

    ('Design / Spec', 'DSL & SPEC authoring',
     'Claude',
     'Generated YAML DSL covering entities, state machine, approval chains, AI agent config, validation rules, and API surface in a single declarative file.',
     'specs/00_main.yaml'),

    ('Architecture', 'Application & technical architecture diagrams',
     'Claude',
     'Asked Claude to produce 6 Mermaid diagrams (layered architecture, deployment, sequence, state machine, AI agent topology, C4 container).',
     'diagrams/architecture.md'),

    ('Data Design', 'Logical and physical data model',
     'Claude',
     'Derived entities from FRD, generated DDL for all tables with constraints and indexes, included migration notes SQLite -> PostgreSQL.',
     'db/database_design.md'),

    ('Implementation', 'Workflow engine, API routes, UI pages',
     'Claude',
     'Generated full Next.js app: state machine, 18 REST endpoints, 6 React pages, Tailwind UI. Iteratively reviewed for SoD and edge cases.',
     'app/* (all source)'),

    ('Implementation', 'AI agent orchestration layer',
     'Claude',
     'Designed a 4-agent orchestrator (Justification, Risk Scoring, Context Summary, Policy) with PII redaction, response caching, structured JSON outputs, and graceful fallback.',
     'app/src/lib/ai-agents.js'),

    ('Runtime AI', 'Justification drafting in production',
     'OpenAI gpt-4o-mini (temp 0.4)',
     'Called server-side with deviation context to draft 100-200 word compliance-friendly justifications.',
     'Live: ai_invocations table'),

    ('Runtime AI', 'Risk scoring in production',
     'OpenAI gpt-4o-mini (temp 0.2, JSON mode)',
     'Computes LOW/MEDIUM/HIGH plus weighted factors and confidence for every submission.',
     'Live: deviation_requests.ai_risk_score, ai_risk_factors'),

    ('Runtime AI', 'Context summarization for approvers',
     'OpenAI gpt-4o-mini (temp 0.3)',
     'Generates one-paragraph customer+service brief; PII-redacted; cached per-deviation.',
     'Live: deviation_requests.ai_context_summary'),

    ('Runtime AI', 'Policy interpretation',
     'OpenAI gpt-4o-mini (temp 0.1, JSON mode)',
     'Filters and ranks the policy library by relevance to the deviation; returns citations with snippets.',
     'Live: deviation_requests.ai_policy_citations'),

    ('Testing', 'Test case generation',
     'Claude',
     'Derived 50+ functional test cases from FRs and user stories; each maps back to requirement IDs.',
     'tests/test_scripts_and_traceability.xlsx (this sheet)'),

    ('Documentation', 'User manual & prompt log',
     'Claude',
     'Authored end-user manual HTML, prompt log document, AI-tool usage matrix (this sheet).',
     'manual/, prompts/'),

    ('QA / Review', 'Traceability matrix',
     'Claude',
     'Walked every FR + NFR and mapped to spec, code path, and test case IDs.',
     'tests/test_scripts_and_traceability.xlsx (Traceability sheet)'),
]

ws3.append(['SDLC Phase', 'Task', 'AI Tool / Model', 'How AI Was Used', 'Output Artifact'])
for row in ai_usage:
    ws3.append(list(row))

widths3 = {'A': 16, 'B': 36, 'C': 28, 'D': 60, 'E': 36}
for col, w in widths3.items():
    ws3.column_dimensions[col].width = w

apply_header(ws3)
apply_data_style(ws3)
ws3.freeze_panes = 'A2'

# Save
out = '/home/claude/sda-workflow/tests/test_scripts_and_traceability.xlsx'
wb.save(out)
print(f"Saved: {out}")
print(f"Test cases: {len(test_cases)}")
print(f"Traceability rows: {len(matrix)}")
print(f"AI usage rows: {len(ai_usage)}")
